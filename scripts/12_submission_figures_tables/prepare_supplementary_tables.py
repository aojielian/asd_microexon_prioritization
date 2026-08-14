#!/usr/bin/env python3
"""Supplementary tables workbook (see ANALYSIS_WORKFLOW.md).

Builds 03_supplementary_tables/Supplementary_Tables_FINAL.xlsx, the
journal-upload version of Tables S1-S11: publication-facing sheet and
    column names only, exact numeric values, no provenance / path /
    workflow fields.  Content sources are the accepted supplementary
    tables (S1-S6) and the manuscript-ready assets (S7-S11); all cleaning
    is a scripted remap (renames/drops of labels only — NO value is
    recomputed).

Also writes:
  03_supplementary_tables/Supplementary_Tables_README.md
"""
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from supplementary_common import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# --------------------------------------------------------------- sources
S1 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S1.tsv")
S2 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S2.tsv")
S3 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S3.tsv")
S4 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S4.tsv")
S5 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S5.tsv")
S6 = os.path.join(D35_SUPP_TAB, "Supplementary_Table_S6.tsv")
S7 = os.path.join(D36_ADJ, "TIER_A_ADJUSTED_TRANSCRIPT_USAGE_M0_M4_M4C.tsv")
S8A = os.path.join(D36_PROT, "TIER_A_CODING_CONSEQUENCE.tsv")
S8B = os.path.join(D36_PROT, "TIER_A_PROTEIN_FEATURES.tsv")
S8C = os.path.join(D36_PROT, "TIER_A_UNIPROT_INTERPRO_MAPPING.tsv")
S8D = os.path.join(D36_PROT, "TIER_A_ALPHAFOLD_CONTEXT.tsv")
S9A = os.path.join(D36_NM, "M4C_NEURON_MERGED_ALL19.tsv")
S9B = os.path.join(D36_NM, "M4C_NEURON_MERGED_TIERA_SUMMARY.tsv")
S9C = os.path.join(D36_NM, "M4C_NEURON_MERGED_VS_PRIMARY_M4C.tsv")
S9D = os.path.join(D36_NM, "NEURON_MERGED_CLR_PC_SUMMARY.tsv")
S9E = os.path.join(D36_NM, "NEURON_MERGED_FRACTIONS.tsv")
S10A = os.path.join(D36_D2, "D2_REPRESENTATIVE_TRANSCRIPTS_ALL19.tsv")
S10B = os.path.join(D36_D2, "D2_REPRESENTATIVE_TRANSCRIPTS_TIERA.tsv")
S11A = os.path.join(D36_ANC, "ANCESTRY_RAW_LEVELS.tsv")
S11B = os.path.join(D36_ANC, "ANCESTRY_MODEL_TERMS.tsv")
S11C = os.path.join(D36_ANC, "ANCESTRY_PUBLIC_VS_PRIMARY_COMPARISON.tsv")

# ------------------------------------------------- scripted cleaning maps
# Exact-string value replacements (internal labels -> publication labels);
# numerics are never touched.
GLOBAL_VALUE_RENAMES = {
    "CONCORDANT_ALL_3_REGIONS": "CONCORDANT_ALL_3_REGIONS",
    "RECIPROCAL_LIFTOVER_CONCORDANT": "RECIPROCAL_LIFTOVER_CONCORDANT",
    "M0_primary": "M0_primary",
    "parametric_simulation_fixed_effect_vcov_B1000_seed42":
        "Parametric simulation of the fixed-effect variance-covariance "
        "matrix (1,000 draws, seed 42)",
}

COL_RENAMES = {
    "source_chr_hg19": "hg19_chr",
    "source_start_hg19": "hg19_start",
    "source_end_hg19": "hg19_end",
    "developmental_timing_tier": "developmental_timing_tier",
    "manuscript_claim_level": "reporting_level",
    "KR_P": "P_Kenward_Roger",
    "P_KR": "P_Kenward_Roger",
    "KR_BH_FDR": "BH_FDR_KR",
    "adjusted_control_usage": "adjusted_control_transcript_usage",
    "adjusted_ASD_usage": "adjusted_ASD_transcript_usage",
    "adjusted_ASD_minus_control":
        "adjusted_difference_ASD_minus_control",
    "dir32_protein_coding_status": "protein_coding_status",
    "microexon_reference_hg38": "microexon_coordinates_hg38",
    "microexon_reference_length_nt": "microexon_length_nt",
    "vs_gencode_boundary_shift_bp": "boundary_shift_vs_gencode_bp",
    "M4_beta_reference": "M4_beta_primary",
    "M4_direction_reference": "M4_direction_primary",
    "M4C_primary_beta_reference": "M4C_primary_beta",
    "M4C_primary_KR_BH_FDR_reference": "M4C_primary_KR_BH_FDR",
    "source_variable": "variable",
    "M4": "analysis_M4",
    "public_datMeta_model": "public_metadata_model",
}

COL_DROPS = {"prediction_method", "direction_check", "source_model_file",
             "source_data_file", "source_file"}

S11_VALUE_RENAMES = {
    "level_order_reference": "level_order_analysis",
    "reference_level_reference": "reference_level_analysis",
    "primary19 M4/M4C reference level = EUR (first factor level, R default "
    "treatment contrasts)":
        "M4/M4C reference level = EUR (first factor level, R default "
        "treatment contrasts)",
    "532 primary19 analysis samples are a subset of the public table "
    "samples":
        "the 532 analysis samples are a subset of the public table "
        "samples",
    "primary19 cortical analysis samples (11 regions, 80 donors)":
        "cortical analysis samples (11 regions, 80 donors)",
    "NO: public table has 808 samples vs 532 primary19 cortical samples":
        "NO: public table has 808 samples vs 532 cortical analysis "
        "samples",
    "100% diagonal agreement across all 80 donors (checked)":
        "100% diagonal agreement across all 80 donors",
}

# ------------------------------------------------------------- titles
TITLES = {
    1: "Microexon event set and cross-species mapping: complete evidence "
       "for the 19 ASD-associated microexon events",
    2: "PsychENCODE diagnosis coefficients under the primary model (M0) "
       "and technical-covariate sensitivity models (M1-M4)",
    3: "PsychENCODE diagnosis coefficients under transcript-set "
       "definitions D0-D3",
    4: "Leave-one-donor-out (LODO) influence summary for all 19 events "
       "under M0 and M4",
    5: "Cell-composition estimates for 532 PsychENCODE cortical samples",
    6: "Composition-adjusted (M4C) event-level results",
    7: "Tier A model-adjusted transcript-usage effects on the probability "
       "scale (M0/M4/M4C)",
    8: "Tier A microexon coding consequences and protein annotation",
    9: "Neuron-merged cell-composition sensitivity results",
    10: "Deterministic representative transcript pair selection (D2) for "
        "all 19 events",
    11: "Reported ancestry categories and model encoding for the "
        "PsychENCODE analysis",
}

INT_RE = re.compile(r"^[+-]?\d+$")
FLT_RE = re.compile(r"^[+-]?(\d+\.\d*|\.\d+|\d+)([eE][+-]?\d+)?$")


def cell(v):
    if v is None or v == "":
        return None
    if INT_RE.match(v):
        return int(v)
    if FLT_RE.match(v):
        return float(v)
    return v


def clean_val(v, extra=None):
    if v is None:
        return None
    if extra and v in extra:
        v = extra[v]
    if v in GLOBAL_VALUE_RENAMES:
        v = GLOBAL_VALUE_RENAMES[v]
    return v


def read_block(path, renames=None, drops=None, extra_vals=None):
    """Return (header, rows) with cleaning applied."""
    rows = rd(path)
    hdr = list(rows[0].keys())
    hdr = [h for h in hdr if h not in (drops or set())]
    hdr = [COL_RENAMES.get(h, h) if (renames is None or renames) else h
           for h in hdr]
    out = []
    for r in rows:
        out.append([clean_val(r_orig, extra_vals) for r_orig in
                    [r[k] for k in list(rows[0].keys())
                     if k not in (drops or set())]])
    return hdr, out


def split_sections(path):
    secs = {}
    order = []
    hdr, rows, name = None, [], None
    for line in open(path):
        line = line.rstrip("\n")
        if line.startswith("## section:"):
            if hdr is not None:
                secs[name] = (hdr, rows)
            name = line.split(":", 1)[1].strip()
            order.append(name)
            hdr, rows = None, []
            continue
        if not line.strip():
            continue
        parts = line.split("\t")
        if hdr is None:
            hdr = parts
        else:
            rows.append(parts)
    if hdr is not None:
        secs[name] = (hdr, rows)
    return order, secs


# ------------------------------------------------------------- sheet plan
def build_sheets():
    """Yield (sheet_name, header, rows, content_note) in workbook order."""
    plan = []

    h, r = read_block(S1)
    assert len(r) == 19
    plan.append(("Table S1", h, r,
                 "Complete per-event evidence for the 19 ASD-associated "
                 "microexon events"))

    h, r = read_block(S2, drops={"gene"})   # upstream column is fully empty
    assert len(r) == 95
    plan.append(("Table S2", h, r,
                 "95 rows: 19 events x 5 models (M0 primary; M1-M4 "
                 "sensitivity)"))

    h, r = read_block(S3)
    assert len(r) == 71      # 19 D0 + 15 D1 + 19 D2 + 18 D3
    plan.append(("Table S3", h, r,
                 "19 events x 4 transcript-set definitions (D0-D3)"))

    h, r = read_block(S4)
    assert len(r) == 38
    plan.append(("Table S4", h, r,
                 "38 rows: 19 events x 2 models (M0, M4)"))

    order, secs = split_sections(S5)
    assert order == ["coverage and reference", "7 broad-cell fractions",
                     "fraction QC", "marker validation",
                     "composition PC scores", "PC loadings",
                     "variance explained"], order
    h, r = secs["coverage and reference"]
    plan.append(("Table S5 coverage and reference", h, r,
                 "Analysis coverage and reference descriptors"))
    h, r = secs["7 broad-cell fractions"]
    assert len(r) == 532
    plan.append(("Table S5 fractions", h, r,
                 "Seven-class cell-composition fractions for 532 cortical "
                 "samples"))
    h, r = secs["fraction QC"]
    assert len(r) == 532
    plan.append(("Table S5 fraction fit quality", h, r,
                 "Fraction-sum and fit diagnostics per sample"))
    h, r = secs["marker validation"]
    plan.append(("Table S5 marker validation", h, r,
                 "Marker-gene validation of the composition classes"))
    h, r = secs["composition PC scores"]
    assert len(r) == 532
    plan.append(("Table S5 PC scores", h, r,
                 "Sample scores on the two composition principal "
                 "components"))
    lh, lr = secs["PC loadings"]
    vh, vr = secs["variance explained"]
    assert len(lr) == 7 and len(vr) == 8   # 6 PCs + retained_k + cumulative
    plan.append(("Table S5 loadings and variance",
                 lh, lr + [[""] * len(lh)] + [vh] + vr,
                 "Class loadings and variance explained for the two "
                 "composition principal components"))

    order, secs = split_sections(S6)
    assert len(order) == 5, order
    names6 = ["Table S6 M4 vs M4C", "Table S6 Tier A exact values",
              "Table S6 Tier A M4C LODO", "Table S6 M4C D0-D3 events",
              "Table S6 M4C D0-D3 set summary"]
    notes6 = ["M4 versus composition-adjusted M4C for all 19 events",
              "Tier A exact M4C estimates (full sample)",
              "Tier A M4C leave-one-donor-out summary",
              "M4C results under transcript-set definitions D0-D3",
              "Set-level direction counts and exact binomial P"]
    for nm, note, sec in zip(names6, notes6, order):
        h, r = secs[sec]
        plan.append((nm, h, r, note))
    assert len(plan[-5][2]) == 19 and len(plan[-4][2]) == 4 and \
        len(plan[-3][2]) == 4

    h, r = read_block(S7, drops=COL_DROPS)
    assert len(r) == 12
    plan.append(("Table S7", h, r,
                 "12 rows: 4 Tier A events x 3 models (M0/M4/M4C) on the "
                 "probability scale"))

    h, r = read_block(S8A)
    assert len(r) == 29
    plan.append(("Table S8 coding consequences", h, r,
                 "Transcript-level coding consequences of microexon "
                 "inclusion (29 transcripts)"))
    h, r = read_block(S8B)
    assert len(r) == 139
    plan.append(("Table S8 protein features", h, r,
                 "UniProt features near the four Tier A insertion sites"))
    h, r = read_block(S8C)
    assert len(r) == 75
    plan.append(("Table S8 UniProt-InterPro map", h, r,
                 "UniProt-InterPro cross-references for the four Tier A "
                 "proteins"))
    h, r = read_block(S8D)
    assert len(r) == 4
    plan.append(("Table S8 AlphaFold context", h, r,
                 "AlphaFold pLDDT context at the four insertion sites"))

    h, r = read_block(S9A)
    assert len(r) == 19
    plan.append(("Table S9 all-event results", h, r,
                 "Neuron-merged M4C results for all 19 events with the "
                 "model formula"))
    h, r = read_block(S9B)
    assert len(r) == 4
    plan.append(("Table S9 Tier A summary", h, r,
                 "Tier A summary under the neuron-merged composition"))
    h, r = read_block(S9C)
    assert len(r) == 19
    plan.append(("Table S9 vs primary M4C", h, r,
                 "Neuron-merged M4C compared with the primary M4C"))
    h, r = read_block(S9D)
    assert len(r) == 6
    plan.append(("Table S9 composition PC summary", h, r,
                 "Neuron-merged six-class composition principal-component "
                 "summary"))
    h, r = read_block(S9E)
    assert len(r) == 532
    plan.append(("Table S9 sample fractions", h, r,
                 "Neuron-merged fractions for 532 samples"))

    h, r = read_block(S10A, drops=COL_DROPS)
    assert len(r) == 19
    plan.append(("Table S10 all events", h, r,
                 "Deterministic D2 representative transcript pairs for all "
                 "19 events"))
    h, r = read_block(S10B, drops=COL_DROPS)
    assert len(r) == 4
    plan.append(("Table S10 Tier A events", h, r,
                 "Deterministic D2 representative transcript pairs for the "
                 "four Tier A events"))

    h, r = read_block(S11A)
    assert len(r) == 5
    plan.append(("Table S11 ancestry levels", h, r,
                 "Five reported ancestry categories with donor and sample "
                 "counts"))
    h, r = read_block(S11B)
    assert len(r) == 10
    plan.append(("Table S11 model encoding", h, r,
                 "Model encoding of the ancestry factor"))
    h, r = read_block(S11C, extra_vals=S11_VALUE_RENAMES)
    assert len(r) == 14
    plan.append(("Table S11 metadata comparison", h, r,
                 "Analysis-model ancestry encoding compared with the "
                 "public metadata table"))

    return plan



def main():
    plan = build_sheets()
    wb = Workbook()
    wb.remove(wb.active)
    toc = wb.create_sheet("Contents")
    toc.append(["Table", "Title", "Sheet", "Sheet content"])
    for c in toc[1]:
        c.font = Font(bold=True)
    seen_table = {}
    for sheet, hdr, rows, note in plan:
        n = int(sheet.split()[1].replace("S", "").split("-")[0]
                if "-" in sheet else sheet.split()[1][1:])
        if sheet.startswith("Table S"):
            n = int(re.match(r"Table S(\d+)", sheet).group(1))
        if n not in seen_table:
            seen_table[n] = TITLES[n]
        assert len(sheet) <= 31, sheet
        ws = wb.create_sheet(sheet)
        ws.append(hdr)
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append([cell(clean_val(v)) if isinstance(v, str) else v
                       for v in r])
        toc.append(["Table S%d" % n, TITLES[n], sheet, note])
    xlsx = os.path.join(TAB_DIR, "Supplementary_Tables_FINAL.xlsx")
    wb.save(xlsx)
    print("wrote Supplementary_Tables_FINAL.xlsx with %d sheets"
          % (len(wb.sheetnames)))

    # ------------------------------------------------------------ README
    lines = [
        "# Supplementary Tables — Molecular Autism supplementary package",
        "",
        "This workbook (Supplementary_Tables_FINAL.xlsx) contains the "
        "final supplementary tables S1-S11. Each table occupies one or "
        "more sheets; the Contents sheet lists every sheet with a short "
        "description. Numeric values are final analysis values and must "
        "be cited exactly as printed.",
        "",
        "## Tables and sheets",
        ""]
    for n in sorted(TITLES):
        lines.append("### Table S%d. %s" % (n, TITLES[n]))
        for sheet, hdr, rows, note in plan:
            if re.match(r"Table S%d( |$)" % n, sheet):
                lines.append("- **%s** (%d data rows): %s"
                             % (sheet, len(rows), note))
        lines.append("")
    lines += [
        "## Abbreviations and notation",
        "",
        "- HsaEX/MmuEX: human/mouse microexon event identifiers "
        "(VastDB-style).",
        "- ASD: autism spectrum disorder diagnosis; coefficients refer "
        "to the ASD diagnosis term in mixed-effects models for "
        "logit-transformed transcript usage.",
        "- M0: clinical-covariate model; M1-M4: technical-covariate "
        "sensitivity models; M4C: M4 plus composition principal "
        "components; neuron-merged M4C: M4C with excitatory and "
        "inhibitory neuron classes merged (sensitivity).",
        "- KR: Kenward-Roger; BH: Benjamini-Hochberg; FDR: false "
        "discovery rate; LRT: likelihood-ratio test; LODO: "
        "leave-one-donor-out.",
        "- D0-D3: transcript-set definitions used in the "
        "definition-sensitivity analysis; D2 is the representative-pair "
        "definition (Table S10).",
        "- PSI: percent-spliced-in; PSI values describe developmental "
        "trajectory context only.",
        "- CompPC/NMCompPC: composition principal components (seven-"
        "class and neuron-merged composition, respectively).",
        "- MANE: Matched Annotation from NCBI and EMBL-EBI; pLDDT: "
        "AlphaFold per-residue confidence (predicted).",
        "- Tier A-D: final evidence tiers; Tier A = cross-cohort "
        "significance (KR FDR < 0.05) with direction concordance.",
        "",
        "All event-level statistics in Tables S1-S6 are identical to the "
        "values reported in the main text and figure legends.",
        ""]
    with open(os.path.join(TAB_DIR, "Supplementary_Tables_README.md"),
              "w") as f:
        f.write("\n".join(lines))
    print("wrote Supplementary_Tables_README.md")
    if clean != "OK":
        raise SystemExit("submission-cleanliness scan ERROR")
    print("table workbook done")


if __name__ == "__main__":
    main()
