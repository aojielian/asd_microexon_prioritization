import os
"""Final Table S1 publication cleanup.

Applies the final publication-clean edits to Table S1 (column removal and reader-facing enum conversions).
Paths: configured via environment variables PROJECT_ROOT, DATA_ROOT, REFERENCE_ROOT, LIFTOVER_PATH (see config/paths_template.yaml).
"""
import shutil, sys

SRC = os.path.join(os.environ.get("PROJECT_ROOT", "."), "41_submission_figures_and_tables/03_supplementary_tables/Supplementary_Tables_Molecular_Autism_SUBMISSION_CLEAN_v2_20260813.xlsx")
DST = os.path.join(os.environ.get("PROJECT_ROOT", "."), "41_submission_figures_and_tables/03_supplementary_tables/Supplementary_Tables_Molecular_Autism_FINAL_SUBMISSION_20260813.xlsx")
S1 = "Table S1"
DROP_HEADER = "developmental_timing_tier"

import openpyxl

# ---------------- section-2 value maps (header -> {old: new}) -----------------
VALUE_MAPS = {
    "liftover_status": {"CONCORDANT_ALL_3_REGIONS": "Concordant"},
    "roundtrip_status": {"RECIPROCAL_LIFTOVER_CONCORDANT": "Reciprocal liftOver concordant"},
    "GENCODE_v33_local_structure_status": {"EQUIVALENT_0_1BP_GRCH38_LOCAL_STRUCTURE": "Equivalent local structure (0-1 bp)"},
    "CHyMErA_direction": {"MICROEXON_INCLUSION_LOSS": "Microexon inclusion loss"},
    "CHyMErA_direction_concordant": {"YES": "Yes", "NO": "No"},
    "Parikshak_direction": {"DOWN_IN_ASD": "Down in ASD", "UP_IN_ASD": "Up in ASD"},
    "developmental_dynamic_status": {"DYNAMIC": "Dynamic", "NON_DYNAMIC": "Non-dynamic"},
    "developmental_trajectory": {"NON_DYNAMIC": "Non-dynamic"},
    "network_module_or_pathway": {"SET_LEVEL_NETWORK_MEMBERSHIP": "Member of curated host-gene network"},
    "GSE30573_direction_concordant": {"CONCORDANT": "Concordant", "DISCORDANT": "Discordant"},
    "PsychENCODE_direction": {"DOWN_IN_ASD": "Down in ASD", "UP_IN_ASD": "Up in ASD"},
    "direction_concordant": {"TRUE": "Yes", "FALSE": "No"},
    "transcript_usage_definition": {"RSEM_effLen_normalized_ratio": "RSEM effective-length-normalized ratio"},
    "negative_evidence_summary": {"GSE30573 small-n context only": "GSE30573 limited context (3 mapped events)"},
}

# ---------------- load v2 reference -------------------------------------------
ref = openpyxl.load_workbook(SRC)
ref_s1 = ref[S1]
ref_rows = list(ref_s1.iter_rows(values_only=True))
ref_header = [str(v) if v is not None else "" for v in ref_rows[0]]
assert DROP_HEADER in ref_header
drop_idx = ref_header.index(DROP_HEADER)  # 0-based col index to delete
n_cols = len(ref_header)
assert n_cols == 53, n_cols

# build expected final S1 grid: v2 minus drop column + maps applied
def norm(v):
    return None if v is None else str(v)

exp_header = [h for i, h in enumerate(ref_header) if i != drop_idx]
exp_grid = []
for r in ref_rows[1:]:
    row = []
    for i, v in enumerate(r):
        if i == drop_idx:
            continue
        h = ref_header[i]
        mp = VALUE_MAPS.get(h, {})
        row.append(mp.get(norm(v), v))
    exp_grid.append(row)
print("expected final cols:", len(exp_header), "rows:", len(exp_grid))

# ---------------- copy + edit ---------------------------------------------------
shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)
ws = wb[S1]
ws.delete_cols(drop_idx + 1, 1)  # openpyxl 1-based

# apply maps by header
hdr = [str(c.value) if c.value is not None else "" for c in ws[1]]
hdr_idx = {h: i + 1 for i, h in enumerate(hdr)}  # 1-based col
assert len(hdr) == 52
applied = 0
for h, mp in VALUE_MAPS.items():
    col = hdr_idx[h]
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        v = norm(cell.value)
        if v in mp:
            cell.value = mp[v]
            applied += 1
        else:
            # keep-value check: expected grid must carry the same value
            er = exp_grid[r - 2][hdr_idx[h] - 1]
            assert norm(er) == norm(cell.value), (h, r, cell.coordinate, v, er)
print("applied:", applied)

# ---------------- full-grid equality check (S1) ----------------------------------
bad = []
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        got = norm(ws.cell(row=r, column=c).value)
        if r == 1:
            exp = exp_header[c - 1]
        else:
            exp = norm(exp_grid[r - 2][c - 1])
        if norm(got) != norm(exp):
            bad.append((r, c, got, exp))
print("S1 grid diffs vs expected:", bad if bad else "NONE")

# ---------------- S2-S12 value-identity + numeric immutability ------------------
s2bad, numb = [], 0
for nm in ref.sheetnames:
    if nm == S1:
        continue
    rws, wvs = ref[nm], wb[nm]
    for row_ref, row_new in zip(rws.iter_rows(), wvs.iter_rows()):
        for c_ref, c_new in zip(row_ref, row_new):
            if norm(c_ref.value) != norm(c_new.value):
                s2bad.append((nm, c_ref.coordinate, norm(c_ref.value), norm(c_new.value)))
            if isinstance(c_ref.value, (int, float)) and c_ref.value != c_new.value:
                numb += 1
# S1 numeric cells: none changed (only text maps) - compare numeric cells directly
s1num = 0
for r in range(1, ref_s1.max_row + 1):
    for c in range(1, ref_s1.max_column + 1):
        v = ref_s1.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            # find same value in final grid (column shifted if after drop)
            nc = c if c <= drop_idx else c - 1
            w = ws.cell(row=r, column=nc).value
            if v != w:
                s1num += 1
print("S2-S12 diffs:", s2bad if s2bad else "NONE")
print("numeric changes S2-S12:", numb, " S1:", s1num)

# ---------------- style spot check ----------------------------------------------
h_bold = all(ws.cell(row=1, column=c).font.bold for c in (1, 13, 24, 35, 46, 52))
d_bold = not ws.cell(row=2, column=1).font.bold
print("header bold row 1:", h_bold, " data not bold:", d_bold)

if bad or s2bad or numb or s1num or not h_bold or not d_bold:
    print("FINAL_CLEAN_ERROR")
    sys.exit(1)

wb.save(DST)
print("wrote", DST)
print("FINAL_S1_CLEAN_OK")
